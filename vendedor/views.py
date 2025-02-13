# vendedor/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill, NamedStyle
)
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import DateAxis
import openpyxl
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from collections import defaultdict
from .models import Vendedor
from pedidos.models import Pedido, PedidoItem, Referencia
from .forms import VendedorForm
from django.contrib.auth.decorators import login_required
from openpyxl.utils import get_column_letter
from functools import wraps
import csv
import datetime
import calendar

def permission_required(min_level):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            # Verifica se o usuário está autenticado e possui o nível mínimo
            if not request.user.is_authenticated or request.user.permission_level < min_level:
                return HttpResponseForbidden("Você não tem permissão para acessar esta página.")
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

# ---------------------------------------------------------------------
# Outras views (lista_vendedores, criar, editar, deletar)
# ---------------------------------------------------------------------

@permission_required(1)
def lista_vendedores(request):
    vendedores_list = Vendedor.objects.all().order_by('codigo')
    
    # Recupera o número da página atual (caso não exista, assume 1)
    page = request.GET.get('page', 1)
    
    # Define quantos itens por página (por exemplo, 10)
    paginator = Paginator(vendedores_list, 10)
    
    try:
        vendedores = paginator.page(page)
    except PageNotAnInteger:
        vendedores = paginator.page(1)
    except EmptyPage:
        vendedores = paginator.page(paginator.num_pages)

    return render(request, 'vendedor/lista_vendedores.html', {
        'vendedores': vendedores
    })


@permission_required(4)
def criar_vendedor(request):
    if request.method == 'POST':
        form = VendedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('vendedor:lista_vendedores')
    else:
        form = VendedorForm()
    return render(request, 'vendedor/form_vendedor.html', {'form': form})


@permission_required(4)
def editar_vendedor(request, pk):
    vendedor = get_object_or_404(Vendedor, pk=pk)
    if request.method == 'POST':
        form = VendedorForm(request.POST, instance=vendedor)
        if form.is_valid():
            form.save()
            return redirect('vendedor:lista_vendedores')
    else:
        form = VendedorForm(instance=vendedor)
    return render(request, 'vendedor/form_vendedor.html', {'form': form})


@permission_required(4)
def deletar_vendedor(request, pk):
    if request.method == 'POST':
        vendedor = get_object_or_404(Vendedor, pk=pk)
        vendedor.delete()
        return redirect('vendedor:lista_vendedores')
    else:
        return redirect('vendedor:lista_vendedores')


@login_required
def download_excel_report(request, vendedor_id):
    """
    Gera um relatório Excel detalhado para cada mês que possua pedidos do vendedor,
    incluindo estatísticas (melhor dia, maior venda, etc.) e um gráfico de barras
    de vendas por dia. Nesta versão:
      - O campo que antes usava "mat_balancinho" é substituído por "Sintético" (nome do produto).
      - São incluídas as colunas para os itens faltantes.
      - O campo 'status' foi substituído por 'status_balancinho'.
    """
    vendedor = get_object_or_404(Vendedor, pk=vendedor_id)
    ano_atual = datetime.datetime.now().year

    pedidos_todos = (
        Pedido.objects.filter(vendedor=vendedor)
        .prefetch_related('itens', 'itens__produto')
        .order_by('-id')
    )

    wb = openpyxl.Workbook()
    ws_padrao = wb.active
    wb.remove(ws_padrao)  # Remove a planilha padrão

    # -----------------------------
    # Definição de estilos globais
    # -----------------------------
    titulo_fill = GradientFill(stop=("017A39", "01A44D"))
    titulo_font = Font(color="FFFFFF", size=16, bold=True)
    titulo_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="017a39"),
        right=Side(style="thin", color="017a39"),
        top=Side(style="thin", color="017a39"),
        bottom=Side(style="thin", color="017a39")
    )

    cabecalho_fill = PatternFill(start_color="017A39", end_color="01A44D", fill_type="solid")
    cabecalho_font = Font(color="FFFFFF", bold=True)
    cabecalho_alignment = Alignment(horizontal="center", vertical="center")

    # Define estilo zebra para linhas pares
    try:
        wb.remove_named_style("ZebraPar")
    except Exception:
        pass
    zebra_par = NamedStyle(name="ZebraPar")
    zebra_par.fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
    zebra_par.border = thin_border
    wb.add_named_style(zebra_par)

    # Estilo para pedidos cancelados
    cancelado_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    nomes_meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]

    dias_semana_pt = {
        "Monday": "Segunda-feira",
        "Tuesday": "Terça-feira",
        "Wednesday": "Quarta-feira",
        "Thursday": "Quinta-feira",
        "Friday": "Sexta-feira",
        "Saturday": "Sábado",
        "Sunday": "Domingo"
    }

    def criar_planilha_mes(mes):
        nome_mes = nomes_meses[mes - 1] if 1 <= mes <= 12 else f"Mes_{mes}"
        ws = wb.create_sheet(title=f"{nome_mes}_{ano_atual}")
        return ws

    def write_info(ws, row_index, col_start, label, value, merge_range):
        ws.merge_cells(merge_range)
        cell = ws.cell(row=row_index, column=col_start)
        cell.value = f"{label}: {value if value is not None else 'N/A'}"
        cell.font = Font(size=12, bold=True, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_index].height = 25

    # Agrupa os pedidos por mês (usando o campo data)
    pedidos_por_mes = defaultdict(list)
    for p in pedidos_todos:
        pedidos_por_mes[p.data.month].append(p)

    # Importações para o gráfico
    from openpyxl.chart import BarChart, Reference

    # Definindo os headers (19 colunas)
    headers = [
        'Pedido ID', 'Cliente', 'Data', 'Status',
        'Sintético', 'Ref. Balancinho', 'Cor Balancinho',
        'Quantidade', 'Tamanho', 'Tipo serviço', 
        'Ref. Palmilha', 'Palmilha',  'Cor Solado',
        'Marca', 'MM Palmilha', 'Espessura Solado',
        'AutorizadoPor', 'Motivo Cancelamento', 'Obs'
    ]

    # Para cada mês com pedidos
    for mes in range(1, 13):
        if mes not in pedidos_por_mes:
            continue

        ws = criar_planilha_mes(mes)
        pedidos_do_mes = pedidos_por_mes[mes]

        # Estatísticas do mês (desconsiderando pedidos cancelados para as vendas)
        pedidos_analise = [pd for pd in pedidos_do_mes if pd.status_balancinho != 'Cancelado']
        total_vendas = len(pedidos_analise)
        cancelados = sum(1 for pd in pedidos_do_mes if pd.status_balancinho == 'Cancelado')
        vendas_por_dia = defaultdict(int)
        total_itens_mes = 0
        maior_venda = None
        maior_venda_itens = 0

        for pedido in pedidos_do_mes:
            if pedido.status_balancinho != 'Cancelado':
                dia = pedido.data.date()
                soma_itens = sum(it.quantidade for it in pedido.itens.all())
                vendas_por_dia[dia] += soma_itens
                total_itens_mes += soma_itens
                if soma_itens > maior_venda_itens:
                    maior_venda_itens = soma_itens
                    maior_venda = pedido

        if vendas_por_dia:
            best_day = max(vendas_por_dia, key=vendas_por_dia.get)
            best_day_count = vendas_por_dia[best_day]
            data_formatada = best_day.strftime("%d/%m/%Y")
            dia_semana_en = best_day.strftime("%A")
            dia_semana_extenso = dias_semana_pt.get(dia_semana_en, dia_semana_en)
            melhor_dia_str = f"{data_formatada} ({dia_semana_extenso})"
        else:
            melhor_dia_str = "N/A"
            best_day_count = 0

        nome_mes = nomes_meses[mes - 1]
        # Título mesclado
        ws.merge_cells('A1:N3')
        cell_titulo = ws['A1']
        cell_titulo.value = f"RELATÓRIO - {nome_mes.upper()} {ano_atual} - Vendedor: {vendedor.nome.upper()}"
        cell_titulo.font = titulo_font
        cell_titulo.alignment = titulo_alignment
        cell_titulo.fill = titulo_fill
        cell_titulo.border = thin_border
        ws.row_dimensions[1].height = 25

        ws.row_dimensions[4].height = 10  # Espaçamento

        linha_info = 5
        write_info(ws, linha_info, 1, "Loja", vendedor.loja, 'A5:D5')
        write_info(ws, linha_info, 5, "Pedidos Cancelados", cancelados, 'E5:H5')
        write_info(ws, linha_info, 9, "Total de Pares Vendidos", total_itens_mes, 'I5:L5')
        linha_info += 1
        write_info(ws, linha_info, 1, "Total de Vendas", total_vendas, 'A6:D6')
        write_info(ws, linha_info, 5, "Melhor Dia", f"{melhor_dia_str} ({best_day_count} pares)", 'E6:H6')
        if maior_venda:
            info_maior_venda = f"Pedido {maior_venda.pk} - {maior_venda.cliente}, {maior_venda_itens} itens"
        else:
            info_maior_venda = "N/A"
        write_info(ws, linha_info, 9, "Maior Venda", info_maior_venda, 'I6:L6')

        # Cabeçalho da tabela de pedidos
        row_header = 20
        for col_index, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=row_header, column=col_index)
            cell.value = header_text
            cell.font = cabecalho_font
            cell.fill = cabecalho_fill
            cell.alignment = cabecalho_alignment
            cell.border = thin_border
        ws.row_dimensions[row_header].height = 25

        linha_atual = row_header + 1
        for pedido in sorted(pedidos_do_mes, key=lambda x: x.id, reverse=True):
            for item in pedido.itens.all():
                data_str = pedido.data.strftime("%d/%m/%Y %H:%M")
                gerente = str(pedido.gerente_cancelamento) if pedido.gerente_cancelamento else '--'
                cancelado_text = pedido.cancelado if pedido.cancelado else 'Sem Motivo'

                # Construindo a linha com 19 colunas conforme os headers:
                # ['Pedido ID', 'Cliente', 'Data', 'Status',
                #  'Sintético', 'Ref. Balancinho', 'Cor Balancinho',
                #  'Quantidade', 'Tamanho', 'Tipo serviço', 
                #  'Ref. Palmilha', 'Palmilha',  'Cor Solado',
                #  'Marca', 'MM Palmilha', 'Espessura Solado',
                #  'AutorizadoPor', 'Motivo Cancelamento', 'Obs']
                row_data = [
                    pedido.pk,                   # Pedido ID
                    pedido.cliente,              # Cliente
                    data_str,                    # Data
                    pedido.status_balancinho,    # Status (substituído)
                    item.produto.nome,           # Sintético (nome do produto)
                    item.ref_balancinho,         # Ref. Balancinho
                    item.cor,                    # Cor Balancinho
                    item.quantidade,             # Quantidade
                    item.tamanho,                # Tamanho
                    item.tipo_servico,           # Tipo serviço
                    item.ref_palmilha,           # Ref. Palmilha
                    item.mat_palmilha,           # Palmilha
                    item.cor_palmilha,           # Cor Solado
                    item.marca,                  # Marca
                    None,                        # MM Palmilha (campo removido do models)
                    item.espessura,              # Espessura Solado
                    gerente,                     # AutorizadoPor
                    cancelado_text,              # Motivo Cancelamento
                    item.obs                   # Obs
                ]
                for col_index, valor in enumerate(row_data, start=1):
                    cell = ws.cell(row=linha_atual, column=col_index)
                    if isinstance(valor, (int, float)):
                        cell.value = f"{int(valor):,}".replace(",", ".")
                    else:
                        cell.value = valor
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = thin_border
                    if (linha_atual % 2) == 0:
                        cell.style = "ZebraPar"
                    if pedido.status_balancinho == 'Cancelado':
                        cell.fill = cancelado_fill
                ws.row_dimensions[linha_atual].height = 20
                linha_atual += 1

        # Ajuste a largura das colunas com base no conteúdo
        max_row = ws.max_row
        max_col = ws.max_column
        for col in range(2, max_col + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row_ in range(1, max_row + 1):
                val = ws.cell(row=row_, column=col).value
                if val:
                    length = len(str(val))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = max_length + 1
        ws.column_dimensions['A'].width = 10

        # -------------------------
        # Criação do Gráfico de Vendas por Dia
        # -------------------------
        if vendas_por_dia:
            last_day = calendar.monthrange(ano_atual, mes)[1]
            chart_data_start = linha_atual + 2  # Linha onde os dados do gráfico serão escritos
            for day_num in range(1, last_day + 1):
                current_date = datetime.date(ano_atual, mes, day_num)
                qtd = vendas_por_dia.get(current_date, 0)
                ws.cell(row=chart_data_start + day_num - 1, column=1, value=day_num)
                ws.cell(row=chart_data_start + day_num - 1, column=2, value=qtd)

            chart = BarChart()
            chart.title = "Vendas por Dia"
            chart.x_axis.title = "Dia do Mês"
            chart.y_axis.title = "Quantidade de Vendas"
            data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_start + last_day - 1)
            cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_start + last_day - 1)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.legend = None
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None

            # Ajuste do tamanho do gráfico
            max_width = 0
            for col_ in range(1, 15):
                col_letter = get_column_letter(col_)
                col_width = ws.column_dimensions[col_letter].width or 10
                max_width += col_width
            chart.width = max_width * 0.145
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['I'].width = 20
            chart.height = 6

            ws.add_chart(chart, "A8")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Relatorio_{vendedor.nome}_{ano_atual}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
