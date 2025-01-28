# produto/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill,
    NamedStyle
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from collections import defaultdict
import datetime
import calendar

from functools import wraps

from pedidos.models import PedidoItem, Pedido
from vendedor.models import Vendedor  # caso precise exibir info do Vendedor
from .models import Produto
from .forms import ProdutoForm


# ---------------------------------------------------
# DECORATOR DE PERMISSÃO
# ---------------------------------------------------
def permission_required(min_level):
    """
    Verifica se o usuário possui um nível de permissão suficiente.
    Caso contrário, retorna 403 (Forbidden).
    """
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated or request.user.permission_level < min_level:
                return HttpResponseForbidden("Você não tem permissão para acessar esta página.")
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


# ---------------------------------------------------
# LISTAGEM DE PRODUTOS
# ---------------------------------------------------
@permission_required(1)
def lista_produtos(request):
    produtos_list = Produto.objects.all().order_by('nome')
    page = request.GET.get('page', 1)
    paginator = Paginator(produtos_list, 10)  # 10 itens por página

    try:
        produtos = paginator.page(page)
    except PageNotAnInteger:
        produtos = paginator.page(1)
    except EmptyPage:
        produtos = paginator.page(paginator.num_pages)

    return render(request, 'produto/lista_produtos.html', {
        'produtos': produtos
    })


# ---------------------------------------------------
# CRIAÇÃO DE PRODUTO
# ---------------------------------------------------
@permission_required(4)
def criar_produto(request):
    if request.method == 'POST':
        form = ProdutoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produto criado com sucesso!')
            return redirect('produto:lista_produtos')
    else:
        form = ProdutoForm()
    return render(request, 'produto/form_produto.html', {'form': form})


# ---------------------------------------------------
# EDIÇÃO DE PRODUTO
# ---------------------------------------------------
@permission_required(3)
def editar_produto(request, pk):
    produto = get_object_or_404(Produto, pk=pk)
    if request.method == 'POST':
        form = ProdutoForm(request.POST, instance=produto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produto atualizado com sucesso!')
            return redirect('produto:lista_produtos')
    else:
        form = ProdutoForm(instance=produto)
    return render(request, 'produto/form_produto.html', {'form': form})


# ---------------------------------------------------
# EXCLUSÃO DE PRODUTO
# ---------------------------------------------------
@permission_required(4)
def deletar_produto(request, pk):
    if request.method == 'POST':
        produto = get_object_or_404(Produto, pk=pk)
        produto.delete()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'error': 'Método não permitido.'}, status=405)


# ---------------------------------------------------
# DOWNLOAD DE EXCEL PARA UM PRODUTO ESPECÍFICO
# ---------------------------------------------------
@login_required
@permission_required(4)
def download_excel_produto(request, produto_id):
    """
    Gera um relatório em Excel para um produto específico:
      - Cria uma aba (sheet) para cada mês do ano atual com vendas.
      - Lista todos os PedidoItem referentes a esse produto no mês.
      - Gera um gráfico de "vendas diárias" (quantidade vendida por dia).
      - Aplica estilos e layouts semelhantes aos relatórios de vendedores.
      - Título centralizado, demais campos à esquerda, etc.
      - Se não houver vendas para o produto, mantém a sheet padrão com uma mensagem.
    """

    produto = get_object_or_404(Produto, pk=produto_id)
    ano_atual = datetime.datetime.now().year

    # Filtra todos os PedidoItems do ano atual para este produto
    itens_do_produto = (
        PedidoItem.objects
                  .select_related('pedido', 'pedido__vendedor')
                  .filter(produto=produto, pedido__data__year=ano_atual)
                  .order_by('-pedido__id')
    )

    # Agrupa itens por mês (1..12)
    itens_por_mes = defaultdict(list)
    for item in itens_do_produto:
        mes = item.pedido.data.month
        itens_por_mes[mes].append(item)

    # Cria o Workbook e pega a sheet padrão
    wb = openpyxl.Workbook()
    ws_padrao = wb.active

    # Variável para controlar se criamos alguma sheet de mês
    criou_aba_vendas = False

    # ----------------------------------------------------
    # 1) Definição de estilos e fills
    # ----------------------------------------------------
    # Estilo do título
    titulo_fill = GradientFill(stop=("017A39", "01A44D"))  # Fundo gradiente (verde)
    titulo_font = Font(color="FFFFFF", size=16, bold=True)
    titulo_alignment = Alignment(horizontal="center", vertical="center")

    # Bordas finas na cor verde
    thin_border = Border(
        left=Side(style="thin", color="017a39"),
        right=Side(style="thin", color="017a39"),
        top=Side(style="thin", color="017a39"),
        bottom=Side(style="thin", color="017a39")
    )

    # Cabeçalho: fundo verde, texto branco
    cabecalho_fill = PatternFill(start_color="017A39", end_color="01A44D", fill_type="solid")
    cabecalho_font = Font(color="FFFFFF", bold=True)
    cabecalho_alignment = Alignment(horizontal="left", vertical="center")

    # Estilo Zebra
    try:
        wb.remove_named_style("ZebraPar")
    except:
        pass
    zebra_par = NamedStyle(name="ZebraPar")
    zebra_par.fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
    zebra_par.border = thin_border
    zebra_par.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wb.add_named_style(zebra_par)

    # Estilo para pedidos cancelados (vermelho claro)
    cancelado_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Nomes dos meses
    nomes_meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]

    # Dias da semana em pt-BR
    dias_semana_pt = {
        "Monday": "Segunda-feira",
        "Tuesday": "Terça-feira",
        "Wednesday": "Quarta-feira",
        "Thursday": "Quinta-feira",
        "Friday": "Sexta-feira",
        "Saturday": "Sábado",
        "Sunday": "Domingo"
    }

    # ----------------------------------------------------
    # Funções auxiliares
    # ----------------------------------------------------
    def criar_planilha_mes(mes):
        """
        Cria a aba e nomeia como 'Mes_Ano'.
        """
        nome_mes = nomes_meses[mes - 1] if 1 <= mes <= 12 else f"Mes_{mes}"
        ws = wb.create_sheet(title=f"{nome_mes}_{ano_atual}")
        return ws

    def write_info(ws, row_index, col_start, label, value, merge_range):
        """
        Escreve um rótulo e valor em células mescladas, alinhados à esquerda.
        """
        ws.merge_cells(merge_range)
        cell = ws.cell(row=row_index, column=col_start)
        cell.value = f"{label}: {value if value is not None else 'N/A'}"
        cell.font = Font(size=12, bold=True, color="333333")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_index].height = 25

    # ----------------------------------------------------
    # 2) Loop pelos meses para gerar planilhas
    # ----------------------------------------------------
    for mes in range(1, 13):
        # Se não há itens no mês, pula
        if mes not in itens_por_mes:
            continue

        ws = criar_planilha_mes(mes)
        criou_aba_vendas = True  # Criamos uma sheet com dados
        lista_itens = itens_por_mes[mes]

        # ------------------------------------------------
        # Cálculo de estatísticas do mês
        # ------------------------------------------------
        total_quantidades = 0
        cancelados = 0
        vendas_por_dia = defaultdict(int)
        pedido_soma_dict = defaultdict(int)

        for item in lista_itens:
            pedido = item.pedido
            qty = item.quantidade
            total_quantidades += qty

            if pedido.status == 'Cancelado':
                cancelados += 1
            else:
                dia = pedido.data.date()
                vendas_por_dia[dia] += qty
                pedido_soma_dict[pedido.id] += qty

        # Maior pedido
        maior_pedido = None
        maior_pedido_qty = 0
        for pid, soma_qtde in pedido_soma_dict.items():
            if soma_qtde > maior_pedido_qty:
                maior_pedido_qty = soma_qtde
                maior_pedido = pid

        # Melhor dia
        if vendas_por_dia:
            best_day = max(vendas_por_dia, key=vendas_por_dia.get)
            best_day_count = vendas_por_dia[best_day]
            data_formatada = best_day.strftime("%d/%m/%Y")
            dia_semana_en = best_day.strftime("%A")
            dia_semana_extenso = dias_semana_pt.get(dia_semana_en, dia_semana_en)
            melhor_dia_str = f"{data_formatada} ({dia_semana_extenso})"
        else:
            melhor_dia_str = "N/A"
            best_day_count = 0

        # ------------------------------------------------
        # Título (mesclado e centralizado)
        # ------------------------------------------------
        nome_mes = nomes_meses[mes - 1]
        ws.merge_cells('A1:N3')
        cell_titulo = ws['A1']
        cell_titulo.value = (
            f"RELATÓRIO DE VENDAS DO PRODUTO - {produto.nome.upper()} - "
            f"{nome_mes.upper()} {ano_atual}"
        )
        cell_titulo.font = titulo_font
        cell_titulo.alignment = titulo_alignment
        cell_titulo.fill = titulo_fill
        cell_titulo.border = thin_border
        ws.row_dimensions[1].height = 25

        # Um espaçamento antes das infos
        ws.row_dimensions[4].height = 10

        # ------------------------------------------------
        # Informações de Resumo
        # ------------------------------------------------
        linha_info = 5
        write_info(ws, linha_info, 1, "Quantidade Total (mês)", total_quantidades, 'A5:D5')
        write_info(ws, linha_info, 5, "Pedidos Cancelados", cancelados, 'E5:H5')
        linha_info += 1
        write_info(
            ws,
            linha_info,
            5,
            "Melhor Dia",
            f"{melhor_dia_str} ({best_day_count} itens)" if best_day_count else "N/A",
            'E6:H6'
        )

        info_maior_venda = "N/A"
        if maior_pedido:
            pedido_obj = Pedido.objects.filter(pk=maior_pedido).first()
            if pedido_obj:
                info_maior_venda = (
                    f"Pedido #{pedido_obj.pk} - Cliente: {pedido_obj.cliente} "
                    f"({maior_pedido_qty} itens)"
                )
        write_info(ws, linha_info, 1, "Maior Pedido", info_maior_venda, 'A6:D6')

        # ------------------------------------------------
        # Cabeçalho da Tabela de Itens
        # ------------------------------------------------
        headers = [
            'Pedido ID', 'Data', 'Status', 'Cliente',
            'Vendedor', 'Quantidade', 'Tamanho', 'Tipo Serviço',
            'Cor', 'Mat Balancinho', 'Mat Palmilha',
            'Tam. Palmilha', 'Observações'
        ]
        row_header = 21
        for col_index, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=row_header, column=col_index)
            cell.value = header_text
            cell.font = cabecalho_font
            cell.fill = cabecalho_fill
            cell.alignment = cabecalho_alignment
            cell.border = thin_border
        ws.row_dimensions[row_header].height = 25

        # ------------------------------------------------
        # Linhas da Tabela
        # ------------------------------------------------
        linha_atual = row_header + 1
        sorted_items = sorted(lista_itens, key=lambda it: it.pedido.id, reverse=True)

        for item in sorted_items:
            pedido = item.pedido
            data_str = pedido.data.strftime("%d/%m/%Y %H:%M") if pedido.data else ""
            row_data = [
                pedido.pk,
                data_str,
                pedido.status,
                pedido.cliente,
                pedido.vendedor.nome if pedido.vendedor else "",
                item.quantidade,
                item.tamanho,
                item.tipo_servico,
                item.cor,
                item.mat_balancinho,
                item.mat_palmilha,
                item.tamanho_palmilha,
                item.obs
            ]
            for col_index, valor in enumerate(row_data, start=1):
                cell = ws.cell(row=linha_atual, column=col_index)
                cell.value = valor
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = thin_border

                # Linha Zebra (pares)
                if (linha_atual % 2) == 0:
                    cell.style = "ZebraPar"

                # Destacar cancelados
                if pedido.status == 'Cancelado':
                    cell.fill = cancelado_fill

            ws.row_dimensions[linha_atual].height = 20
            linha_atual += 1

        # ------------------------------------------------
        # Ajuste de largura das colunas
        # ------------------------------------------------
        max_row = ws.max_row
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row_ in range(1, max_row + 1):
                val = ws.cell(row=row_, column=col).value
                if val:
                    length = len(str(val))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = max_length + 2

        # ------------------------------------------------
        # Gráfico de Vendas Diárias (BarChart)
        # ------------------------------------------------
        if vendas_por_dia:
            last_day = calendar.monthrange(ano_atual, mes)[1]
            inicio_linha_grafico = linha_atual + 2
            linha_chart = inicio_linha_grafico + 1

            # Preenche (Dia do mês / Quantidade)
            for day_num in range(1, last_day + 1):
                current_date = datetime.date(ano_atual, mes, day_num)
                qtd = vendas_por_dia.get(current_date, 0)
                ws.cell(row=linha_chart, column=1, value=day_num)
                ws.cell(row=linha_chart, column=2, value=qtd)
                linha_chart += 1

            chart = BarChart()
            chart.title = "Vendas Diárias (Quantidade de Itens)"
            chart.x_axis.axId = 10
            chart.y_axis.axId = 20
            chart.x_axis.crossAx = 20
            chart.y_axis.crossAx = 10

            data = Reference(ws, min_col=2, min_row=inicio_linha_grafico + 1, max_row=linha_chart - 1)
            cats = Reference(ws, min_col=1, min_row=inicio_linha_grafico + 1, max_row=linha_chart - 1)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.legend = None

            # Remover linhas de grade
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None

            # Títulos e eixos
            chart.x_axis.title = "Dia do Mês"
            chart.y_axis.title = "Quantidade"

            # Cor das barras (verde)
            if chart.series:
                s = chart.series[0]
                s.graphicalProperties.solidFill = "017a39"
                s.graphicalProperties.line.solidFill = "017a39"

            # Ajuste básico de largura/altura
            max_width = 0
            for col_ in range(1, 16):
                col_letter = get_column_letter(col_)
                col_width = ws.column_dimensions[col_letter].width or 10
                max_width += col_width
            chart.width = max_width * 0.1
            chart.height = 6

            # Coloca o gráfico em A + (linha)
            ws.add_chart(chart, f"A7")

    # ----------------------------------------------------
    # 3) Verifica se criamos pelo menos uma sheet de vendas
    # ----------------------------------------------------
    if criou_aba_vendas:
        # Remove a sheet padrão, pois já existem outras planilhas com dados
        wb.remove(ws_padrao)
    else:
        # Não houve nenhum item de vendas para o produto neste ano
        # Renomeia a sheet padrão e adiciona uma mensagem simples
        ws_padrao.title = "Sem dados"
        ws_padrao['A1'] = "Não há registros de vendas para este produto no ano atual."

    # ----------------------------------------------------
    # 4) Retorna o arquivo Excel como resposta de download
    # ----------------------------------------------------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Relatorio_Produto_{produto.nome}_{ano_atual}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response